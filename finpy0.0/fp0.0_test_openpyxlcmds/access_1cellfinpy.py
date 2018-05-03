from openpyxl import load_workbook
#data_only = True pulls excel output, not the formula
wb = load_workbook('budgetpy0_0.xlsx', data_only = True)
print('\n', wb.sheetnames)

retire_math = wb.active

#test printing out strings from excel, cell = 'Interest'
test_text = retire_math['A2']
print(test_text.value)

#test printing out X% strings, cell = '8%'
test_percent_text = retire_math['B2']
#prints out raw # of %, 8% prints out ass 0.08
print(test_percent_text.value)

#test printing out eq's, cell = '=(C15-B16)*$F$12',
#output val of cell = '$3,026,691.81', cell = C16
test_eqs = retire_math['C16']
#result: this prints out the output see line 2 comment
print('$',test_eqs.value)
