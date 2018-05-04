from openpyxl import load_workbook
wb = load_workbook('budgetpy0.xlsx', data_only = True)
print('\n', wb.sheetnames)

#set vars to sheets to use to change through sheets
month_overview = wb['month_overview']
test_text = month_overview['A1']
print(test_text.value)

w1 = wb['w1']
test_text2 = w1['A1']
print(test_text2.value)

#prints out <Cell 'month_overview'.A8>, etc, not the cell value
# mand_exp_catagories = month_overview['A7':'A9']
# print(mand_exp_catagories)

#print out values of row range
for row in month_overview['A7':'A9']:
    for cell in row:
        print(cell.value)

#try this same thing, jumping over merged cells
for row in month_overview['A7':'A35']:
    for cell in row:
        print(cell.value)
#works, formatting is tough to read though, prob print in chunks
