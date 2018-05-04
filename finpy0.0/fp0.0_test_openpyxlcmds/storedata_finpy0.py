from openpyxl import load_workbook
wb = load_workbook('budgetpy0.xlsx')
print('\n', wb.sheetnames)

#set vars to sheetname
month_overview = wb['Month Overview']

#set value to cell
wife_pay = month_overview['E1']
wife_pay.value = 968
print(wife_pay.value)
wb.save('budgetpy0.xlsx')

#see if it sticks
wife_payinp = month_overview['E1']
print(wife_payinp.value)
#yes this works, reflects in xlsx after close/open

#try dollar value
husband_pay = month_overview['G1']
husband_pay.value = '$2000'
print(husband_pay.value)
wb.save('budgetpy0.xlsx')

#see if dollar value string sticks
hus_payinp = month_overview['G1']
print(hus_payinp.value)
#yes it sticks

#see if it works for the excel hardcoded eq below it
# husband_pay_periods = month_overview['G2']
# husband_pay_periods.value = 2
# print(month_overview['3'].value)
#this doesn't work

#set second part of eq to calculate total pay
husband_pay_periods = month_overview['G2']
husband_pay_periods.value = 3
wb.save('budgetpy0.xlsx')

#print out excel eq output
total_hus_pay = month_overview['G3']
print('$', total_hus_pay.value)

#this works, but editing the excel sheet via this method may have cleared
#all of it's hardcoded formulas

#confirmed it does this: if pre-existing wb is edited, all
#eqs are cleared out once edited
