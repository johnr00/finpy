#import openpyxl, create wb, create sheetnames
from openpyxl import Workbook

wb = Workbook()
wb.save('access_assign_1cell.xlsx')
ws = wb.active
#access 1 cell direct @ location, and call it up basically
c = ws['A4']

#or
d = ws.cell('A5')

#or
e = ws.cell(row = 6, column = 1)
