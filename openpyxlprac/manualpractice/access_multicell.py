#import, create wb, pull active ws, save wb as .xlsx
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
wb.save('access_multicell.xlsx')

#access cells using slicing, 'val':'val'
cell_range = ws['A1':'C2']

#iter through whole file w .row
ws['C9'] = 'hello world'
ws.rows

#iter by columns
ws.columns
