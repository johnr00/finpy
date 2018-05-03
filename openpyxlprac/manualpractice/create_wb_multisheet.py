from openpyxl import Workbook
wb = Workbook()
#remember autocreates 1 sheet, assign it w/ ws = wb.active
wb.save('create_wb_multisheet.xlsx')

#make sheet #1
w1 = wb.create_sheet()
#title sheet #1
w1.title = 'Sheet 1 I made'

#make sheet #2, name sheet #2
w2 = wb.create_sheet()
w2.title = 'Sheet 2 I made'

#make sheet #3, name sheet #3
w3 = wb.create_sheet()
w3.title = 'Sheet 3 I made'
wb.save('create_wb_multisheet.xlsx')

#print wb sheet names
print(wb.sheetnames)
# or loop
for sheet in wb:
    print(sheet)

# or loop
for sheet in wb:
    print(sheet.title)


#pulls you to worksheet 'ws' to write in
ws = wb.active
