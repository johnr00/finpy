from openpyxl import load_workbook

wb = load_workbook('budgetpy_c.xlsx')

print(wb.sheetnames)

ws = wb.active

for cell in ws['Month Overview']:
    print(cell.value)
