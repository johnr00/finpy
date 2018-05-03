from openpyxl import load_workbook

wb = load_workbook('budgetpy_c.xlsx')
print(wb.sheetnames)
