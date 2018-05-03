from openpyxl import Workbook
wb = Workbook()
wb.save('storedata.xlsx')
ws = wb.active

c = ws['A1']

#assign value with .value

c.value = 'hello world in A1'

#print that assigned value. print(c) will just print metadata
print(c.value)
#just wb.save doesn't work, need filename everytime
wb.save('storedata.xlsx')

#enable type and format reference, i.e. %'s,
wb.guess_types = True
c.value = '12%'
print(c.value)
#or
wb.guess_types = True
d = ws['A2']
d.value = '15%'
print(d.value)
