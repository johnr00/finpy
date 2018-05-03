
#Write to a Cell: call a cell, or row/column notation w/ cell()
from openpyxl import workbook

book = Workbook()
sheet = book.active

#method 1
sheet['A1'] = 1
#method 2
sheet.cell(row = 2, column = 2).value = 2

book.save('write2cell.xlsx')

#Appending cells: append() to append (add) group of values at bottom of current
#sheet

from openpyxl import workbook

book = Workbook()
sheet = book.active

#append 3 columns of data into the current sheet
rows = (
    (88,46, 57),
    (89, 38, 12),
    (23, 59, 78),
    (56, 21, 98),
    (24, 18, 43),
    (34, 15, 67),
)

#data is stored in a tuple of tuples... whats a tuple
for row in rows:
    #go row by row and insert data row w/ append()
    sheet.append(row)

book.save('appending.xlsx')
